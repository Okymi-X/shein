# -*- coding: utf-8 -*-
"""
SHEIN_SEN - Générateur de Récapitulatifs
Génération de rapports Excel et PDF pour les commandes groupées
"""

import pandas as pd
import json
from datetime import datetime, timedelta
from typing import Dict, List, Optional, Any
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, Reference
from loguru import logger

from config import Config
from data_manager import DataManager

class RecapExporter:
    """Générateur de récapitulatifs et exports"""
    
    def __init__(self):
        self.data_manager = DataManager()
        self.setup_logging()
    
    def setup_logging(self):
        """Configuration des logs"""
        logger.add(
            f"{Config.LOGS_DIR}/recap_export.log",
            rotation="1 day",
            retention="30 days",
            level="INFO"
        )
    
    def generate_complete_recap(self, output_path: Optional[str] = None) -> str:
        """Générer un récapitulatif complet Excel"""
        try:
            if not output_path:
                timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
                output_path = f"{Config.DATA_DIR}/recap_complet_shein_sen_{timestamp}.xlsx"
            
            # Récupérer toutes les données
            all_orders = self.data_manager.get_all_orders()
            stats = self.data_manager.get_statistics()
            
            if not all_orders:
                logger.warning("Aucune commande trouvée pour le récapitulatif")
                return ""
            
            # Créer le fichier Excel avec plusieurs feuilles
            with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
                # 1. Feuille principale - Toutes les commandes
                self._create_orders_sheet(all_orders, writer)
                
                # 2. Feuille résumé par utilisateur
                self._create_users_summary_sheet(all_orders, writer)
                
                # 3. Feuille statistiques
                self._create_statistics_sheet(stats, writer)
                
                # 4. Feuille produits groupés
                self._create_products_summary_sheet(all_orders, writer)
                
                # 5. Feuille timeline
                self._create_timeline_sheet(all_orders, writer)
            
            # Appliquer le formatage avancé
            self._apply_advanced_formatting(output_path)
            
            logger.info(f"Récapitulatif complet généré: {output_path}")
            return output_path
            
        except Exception as e:
            logger.error(f"Erreur génération récapitulatif: {e}")
            return ""
    
    def _create_orders_sheet(self, orders: List[Dict], writer):
        """Créer la feuille des commandes détaillées"""
        try:
            # Préparer les données
            orders_data = []
            
            for order in orders:
                orders_data.append({
                    'ID Commande': order.get('order_id', ''),
                    'Téléphone': order.get('user_phone', ''),
                    'Nom Utilisateur': order.get('user_name', ''),
                    'URL Produit': order.get('product_url', ''),
                    'Taille': order.get('size', ''),
                    'Couleur': order.get('color', ''),
                    'Quantité': order.get('quantity', 0),
                    'Prix Estimé': order.get('estimated_price', 0),
                    'Statut': order.get('status', ''),
                    'Date Création': self._format_datetime(order.get('created_at')),
                    'Date Traitement': self._format_datetime(order.get('processed_at')),
                    'Notes': order.get('notes', '')
                })
            
            df = pd.DataFrame(orders_data)
            df.to_excel(writer, sheet_name='Commandes_Détaillées', index=False)
            
        except Exception as e:
            logger.error(f"Erreur création feuille commandes: {e}")
    
    def _create_users_summary_sheet(self, orders: List[Dict], writer):
        """Créer la feuille résumé par utilisateur"""
        try:
            # Grouper par utilisateur
            user_summary = {}
            
            for order in orders:
                phone = order.get('user_phone', '')
                if phone not in user_summary:
                    user_summary[phone] = {
                        'Téléphone': phone,
                        'Nom': order.get('user_name', ''),
                        'Total Articles': 0,
                        'Commandes Pending': 0,
                        'Commandes Completed': 0,
                        'Commandes Failed': 0,
                        'Prix Total Estimé': 0,
                        'Première Commande': order.get('created_at', ''),
                        'Dernière Commande': order.get('created_at', '')
                    }
                
                # Mettre à jour les totaux
                user_summary[phone]['Total Articles'] += order.get('quantity', 0)
                user_summary[phone]['Prix Total Estimé'] += order.get('estimated_price', 0)
                
                # Compter par statut
                status = order.get('status', '')
                if status == 'pending':
                    user_summary[phone]['Commandes Pending'] += 1
                elif status == 'completed':
                    user_summary[phone]['Commandes Completed'] += 1
                elif status == 'failed':
                    user_summary[phone]['Commandes Failed'] += 1
                
                # Mettre à jour les dates
                created_at = order.get('created_at', '')
                if created_at > user_summary[phone]['Dernière Commande']:
                    user_summary[phone]['Dernière Commande'] = created_at
                if created_at < user_summary[phone]['Première Commande']:
                    user_summary[phone]['Première Commande'] = created_at
            
            # Formater les dates
            for user_data in user_summary.values():
                user_data['Première Commande'] = self._format_datetime(user_data['Première Commande'])
                user_data['Dernière Commande'] = self._format_datetime(user_data['Dernière Commande'])
            
            df = pd.DataFrame(list(user_summary.values()))
            df.to_excel(writer, sheet_name='Résumé_Utilisateurs', index=False)
            
        except Exception as e:
            logger.error(f"Erreur création feuille utilisateurs: {e}")
    
    def _create_statistics_sheet(self, stats: Dict, writer):
        """Créer la feuille des statistiques"""
        try:
            stats_data = [
                {'Métrique': 'Total Commandes', 'Valeur': stats.get('total_orders', 0)},
                {'Métrique': 'Total Utilisateurs', 'Valeur': stats.get('total_users', 0)},
                {'Métrique': 'Total Articles', 'Valeur': stats.get('total_items', 0)},
                {'Métrique': 'Commandes En Attente', 'Valeur': stats.get('pending_orders', 0)},
                {'Métrique': 'Commandes Complétées', 'Valeur': stats.get('completed_orders', 0)},
                {'Métrique': 'Taux de Réussite (%)', 'Valeur': self._calculate_success_rate(stats)},
                {'Métrique': 'Moyenne Articles/Utilisateur', 'Valeur': self._calculate_avg_items_per_user(stats)},
                {'Métrique': 'Dernière Mise à Jour', 'Valeur': self._format_datetime(stats.get('last_updated'))}
            ]
            
            df = pd.DataFrame(stats_data)
            df.to_excel(writer, sheet_name='Statistiques', index=False)
            
        except Exception as e:
            logger.error(f"Erreur création feuille statistiques: {e}")
    
    def _create_products_summary_sheet(self, orders: List[Dict], writer):
        """Créer la feuille résumé par produit"""
        try:
            # Grouper par URL de produit
            product_summary = {}
            
            for order in orders:
                url = order.get('product_url', '')
                if url not in product_summary:
                    product_summary[url] = {
                        'URL Produit': url,
                        'Total Commandé': 0,
                        'Utilisateurs Différents': set(),
                        'Tailles Demandées': set(),
                        'Couleurs Demandées': set(),
                        'Statut Majoritaire': '',
                        'Prix Total Estimé': 0
                    }
                
                # Mettre à jour les totaux
                product_summary[url]['Total Commandé'] += order.get('quantity', 0)
                product_summary[url]['Prix Total Estimé'] += order.get('estimated_price', 0)
                
                # Ajouter aux ensembles
                if order.get('user_phone'):
                    product_summary[url]['Utilisateurs Différents'].add(order.get('user_phone'))
                if order.get('size'):
                    product_summary[url]['Tailles Demandées'].add(order.get('size'))
                if order.get('color'):
                    product_summary[url]['Couleurs Demandées'].add(order.get('color'))
            
            # Convertir les ensembles en chaînes
            products_data = []
            for url, data in product_summary.items():
                products_data.append({
                    'URL Produit': url[:50] + '...' if len(url) > 50 else url,
                    'Total Commandé': data['Total Commandé'],
                    'Nb Utilisateurs': len(data['Utilisateurs Différents']),
                    'Tailles': ', '.join(sorted(data['Tailles Demandées'])),
                    'Couleurs': ', '.join(sorted(data['Couleurs Demandées'])),
                    'Prix Total Estimé': data['Prix Total Estimé']
                })
            
            # Trier par quantité totale
            products_data.sort(key=lambda x: x['Total Commandé'], reverse=True)
            
            df = pd.DataFrame(products_data)
            df.to_excel(writer, sheet_name='Résumé_Produits', index=False)
            
        except Exception as e:
            logger.error(f"Erreur création feuille produits: {e}")
    
    def _create_timeline_sheet(self, orders: List[Dict], writer):
        """Créer la feuille timeline des commandes"""
        try:
            # Grouper par jour
            daily_stats = {}
            
            for order in orders:
                created_at = order.get('created_at', '')
                if created_at:
                    try:
                        date = datetime.fromisoformat(created_at.replace('Z', '+00:00')).date()
                        date_str = date.strftime('%Y-%m-%d')
                        
                        if date_str not in daily_stats:
                            daily_stats[date_str] = {
                                'Date': date_str,
                                'Nouvelles Commandes': 0,
                                'Total Articles': 0,
                                'Nouveaux Utilisateurs': set()
                            }
                        
                        daily_stats[date_str]['Nouvelles Commandes'] += 1
                        daily_stats[date_str]['Total Articles'] += order.get('quantity', 0)
                        if order.get('user_phone'):
                            daily_stats[date_str]['Nouveaux Utilisateurs'].add(order.get('user_phone'))
                    
                    except Exception:
                        continue
            
            # Convertir en liste et trier
            timeline_data = []
            for date_str, data in sorted(daily_stats.items()):
                timeline_data.append({
                    'Date': date_str,
                    'Nouvelles Commandes': data['Nouvelles Commandes'],
                    'Total Articles': data['Total Articles'],
                    'Nouveaux Utilisateurs': len(data['Nouveaux Utilisateurs'])
                })
            
            df = pd.DataFrame(timeline_data)
            df.to_excel(writer, sheet_name='Timeline', index=False)
            
        except Exception as e:
            logger.error(f"Erreur création feuille timeline: {e}")
    
    def _apply_advanced_formatting(self, file_path: str):
        """Appliquer un formatage avancé au fichier Excel"""
        try:
            wb = openpyxl.load_workbook(file_path)
            
            # Couleurs du thème
            header_color = "366092"
            accent_color = "D9E2F3"
            success_color = "C6EFCE"
            warning_color = "FFEB9C"
            error_color = "FFC7CE"
            
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                
                # Style pour les en-têtes
                header_font = Font(bold=True, color="FFFFFF", size=12)
                header_fill = PatternFill(start_color=header_color, end_color=header_color, fill_type="solid")
                header_alignment = Alignment(horizontal="center", vertical="center")
                
                # Bordures
                thin_border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
                
                # Appliquer le style aux en-têtes
                if ws.max_row > 0:
                    for cell in ws[1]:
                        cell.font = header_font
                        cell.fill = header_fill
                        cell.alignment = header_alignment
                        cell.border = thin_border
                
                # Formatage conditionnel pour la colonne Statut
                if sheet_name == 'Commandes_Détaillées':
                    self._apply_status_formatting(ws, success_color, warning_color, error_color)
                
                # Ajuster la largeur des colonnes
                for column in ws.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    
                    for cell in column:
                        try:
                            if cell.value:
                                cell_length = len(str(cell.value))
                                if cell_length > max_length:
                                    max_length = cell_length
                        except:
                            pass
                    
                    adjusted_width = min(max_length + 2, 60)
                    ws.column_dimensions[column_letter].width = adjusted_width
                
                # Figer la première ligne
                ws.freeze_panes = 'A2'
            
            # Ajouter des graphiques si possible
            self._add_charts(wb)
            
            wb.save(file_path)
            
        except Exception as e:
            logger.error(f"Erreur formatage avancé: {e}")
    
    def _apply_status_formatting(self, ws, success_color: str, warning_color: str, error_color: str):
        """Appliquer un formatage conditionnel pour les statuts"""
        try:
            # Trouver la colonne Statut
            status_col = None
            for col in range(1, ws.max_column + 1):
                if ws.cell(1, col).value == 'Statut':
                    status_col = col
                    break
            
            if status_col:
                for row in range(2, ws.max_row + 1):
                    cell = ws.cell(row, status_col)
                    status = str(cell.value).lower() if cell.value else ''
                    
                    if 'completed' in status or 'complété' in status:
                        cell.fill = PatternFill(start_color=success_color, end_color=success_color, fill_type="solid")
                    elif 'pending' in status or 'attente' in status:
                        cell.fill = PatternFill(start_color=warning_color, end_color=warning_color, fill_type="solid")
                    elif 'failed' in status or 'échec' in status:
                        cell.fill = PatternFill(start_color=error_color, end_color=error_color, fill_type="solid")
            
        except Exception as e:
            logger.error(f"Erreur formatage statuts: {e}")
    
    def _add_charts(self, wb):
        """Ajouter des graphiques aux feuilles appropriées"""
        try:
            # Graphique pour les statistiques
            if 'Statistiques' in wb.sheetnames:
                ws_stats = wb['Statistiques']
                
                # Créer un graphique en barres
                chart = BarChart()
                chart.title = "Statistiques SHEIN_SEN"
                chart.y_axis.title = 'Valeurs'
                chart.x_axis.title = 'Métriques'
                
                # Données pour le graphique (exclure les métriques non numériques)
                data = Reference(ws_stats, min_col=2, min_row=1, max_row=6, max_col=2)
                cats = Reference(ws_stats, min_col=1, min_row=2, max_row=6)
                
                chart.add_data(data, titles_from_data=True)
                chart.set_categories(cats)
                
                # Ajouter le graphique à la feuille
                ws_stats.add_chart(chart, "D2")
            
        except Exception as e:
            logger.error(f"Erreur ajout graphiques: {e}")
    
    def generate_pdf_summary(self, output_path: Optional[str] = None) -> str:
        """Générer un résumé PDF"""
        try:
            from reportlab.lib.pagesizes import A4
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib import colors
            from reportlab.lib.units import inch
            
            if not output_path:
                timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
                output_path = f"{Config.DATA_DIR}/resume_shein_sen_{timestamp}.pdf"
            
            doc = SimpleDocTemplate(output_path, pagesize=A4, topMargin=0.5*inch)
            styles = getSampleStyleSheet()
            story = []
            
            # Style personnalisé pour le titre
            title_style = ParagraphStyle(
                'CustomTitle',
                parent=styles['Heading1'],
                fontSize=24,
                spaceAfter=30,
                alignment=1,  # Centré
                textColor=colors.HexColor('#366092')
            )
            
            # Titre principal
            story.append(Paragraph("SHEIN_SEN", title_style))
            story.append(Paragraph("Système d'Automatisation des Commandes Groupées", styles['Heading2']))
            story.append(Spacer(1, 20))
            
            # Date de génération
            story.append(Paragraph(f"Rapport généré le: {datetime.now().strftime('%d/%m/%Y à %H:%M')}", styles['Normal']))
            story.append(Spacer(1, 30))
            
            # Statistiques globales
            stats = self.data_manager.get_statistics()
            story.append(Paragraph("📊 Statistiques Globales", styles['Heading2']))
            
            stats_data = [
                ['Métrique', 'Valeur'],
                ['Total Commandes', f"{stats.get('total_orders', 0):,}"],
                ['Total Utilisateurs', f"{stats.get('total_users', 0):,}"],
                ['Total Articles', f"{stats.get('total_items', 0):,}"],
                ['Commandes En Attente', f"{stats.get('pending_orders', 0):,}"],
                ['Commandes Complétées', f"{stats.get('completed_orders', 0):,}"],
                ['Taux de Réussite', f"{self._calculate_success_rate(stats):.1f}%"]
            ]
            
            stats_table = Table(stats_data, colWidths=[3*inch, 2*inch])
            stats_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#366092')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 12),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#F0F0F0')),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE')
            ]))
            
            story.append(stats_table)
            story.append(Spacer(1, 30))
            
            # Top utilisateurs
            story.append(Paragraph("👥 Top 5 Utilisateurs", styles['Heading2']))
            top_users = self._get_top_users(5)
            
            if top_users:
                users_data = [['Utilisateur', 'Total Articles', 'Commandes']]
                for user in top_users:
                    users_data.append([
                        user['name'][:20],
                        str(user['total_items']),
                        str(user['total_orders'])
                    ])
                
                users_table = Table(users_data, colWidths=[2.5*inch, 1.5*inch, 1.5*inch])
                users_table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#366092')),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                    ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, 0), 10),
                    ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#F0F0F0')),
                    ('GRID', (0, 0), (-1, -1), 1, colors.black)
                ]))
                
                story.append(users_table)
            
            story.append(Spacer(1, 30))
            
            # Commandes récentes
            story.append(Paragraph("📋 Dernières Commandes", styles['Heading2']))
            recent_orders = self.data_manager.get_all_orders()
            recent_orders = sorted(recent_orders, key=lambda x: x.get('created_at', ''), reverse=True)[:10]
            
            if recent_orders:
                orders_data = [['Date', 'Utilisateur', 'Taille', 'Couleur', 'Qté', 'Statut']]
                
                for order in recent_orders:
                    orders_data.append([
                        self._format_date_short(order.get('created_at')),
                        order.get('user_name', '')[:15],
                        order.get('size', ''),
                        order.get('color', '')[:10],
                        str(order.get('quantity', '')),
                        order.get('status', '')[:8]
                    ])
                
                orders_table = Table(orders_data, colWidths=[1*inch, 1.5*inch, 0.8*inch, 1*inch, 0.7*inch, 1*inch])
                orders_table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#366092')),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                    ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, 0), 9),
                    ('FONTSIZE', (0, 1), (-1, -1), 8),
                    ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#F0F0F0')),
                    ('GRID', (0, 0), (-1, -1), 1, colors.black)
                ]))
                
                story.append(orders_table)
            
            # Pied de page
            story.append(Spacer(1, 50))
            story.append(Paragraph("---", styles['Normal']))
            story.append(Paragraph("SHEIN_SEN - Système d'Automatisation des Commandes Groupées Shein au Sénégal", styles['Normal']))
            story.append(Paragraph("Organise, optimise et facilite les commandes Shein collectives.", styles['Italic']))
            
            # Générer le PDF
            doc.build(story)
            
            logger.info(f"PDF généré: {output_path}")
            return output_path
            
        except Exception as e:
            logger.error(f"Erreur génération PDF: {e}")
            return ""
    
    def _get_top_users(self, limit: int = 5) -> List[Dict]:
        """Obtenir les top utilisateurs par nombre d'articles"""
        try:
            all_orders = self.data_manager.get_all_orders()
            user_stats = {}
            
            for order in all_orders:
                phone = order.get('user_phone', '')
                if phone not in user_stats:
                    user_stats[phone] = {
                        'name': order.get('user_name', ''),
                        'total_items': 0,
                        'total_orders': 0
                    }
                
                user_stats[phone]['total_items'] += order.get('quantity', 0)
                user_stats[phone]['total_orders'] += 1
            
            # Trier par nombre total d'articles
            top_users = sorted(user_stats.values(), key=lambda x: x['total_items'], reverse=True)
            
            return top_users[:limit]
            
        except Exception as e:
            logger.error(f"Erreur récupération top utilisateurs: {e}")
            return []
    
    def _format_datetime(self, dt_str: str) -> str:
        """Formater une chaîne datetime"""
        if not dt_str:
            return ''
        
        try:
            dt = datetime.fromisoformat(dt_str.replace('Z', '+00:00'))
            return dt.strftime('%d/%m/%Y %H:%M')
        except:
            return dt_str
    
    def _format_date_short(self, dt_str: str) -> str:
        """Formater une date en format court"""
        if not dt_str:
            return ''
        
        try:
            dt = datetime.fromisoformat(dt_str.replace('Z', '+00:00'))
            return dt.strftime('%d/%m')
        except:
            return dt_str
    
    def _calculate_success_rate(self, stats: Dict) -> float:
        """Calculer le taux de réussite"""
        total = stats.get('total_orders', 0)
        completed = stats.get('completed_orders', 0)
        
        if total == 0:
            return 0.0
        
        return (completed / total) * 100
    
    def _calculate_avg_items_per_user(self, stats: Dict) -> float:
        """Calculer la moyenne d'articles par utilisateur"""
        total_items = stats.get('total_items', 0)
        total_users = stats.get('total_users', 0)
        
        if total_users == 0:
            return 0.0
        
        return total_items / total_users
    
    def export_for_whatsapp(self, user_phone: str = None) -> str:
        """Générer un résumé formaté pour WhatsApp"""
        try:
            if user_phone:
                # Résumé pour un utilisateur spécifique
                return self.data_manager.generate_user_summary(user_phone)
            else:
                # Résumé global
                stats = self.data_manager.get_statistics()
                
                message_parts = [
                    "📊 *SHEIN_SEN - Résumé Global*",
                    "",
                    f"📦 Total commandes: *{stats.get('total_orders', 0)}*",
                    f"👥 Total utilisateurs: *{stats.get('total_users', 0)}*",
                    f"🛍️ Total articles: *{stats.get('total_items', 0)}*",
                    "",
                    f"⏳ En attente: {stats.get('pending_orders', 0)}",
                    f"✅ Complétées: {stats.get('completed_orders', 0)}",
                    f"📈 Taux réussite: {self._calculate_success_rate(stats):.1f}%",
                    "",
                    f"🕐 Dernière MAJ: {self._format_datetime(stats.get('last_updated'))}",
                    "",
                    "_SHEIN_SEN - Automatisation Commandes Groupées_"
                ]
                
                return "\n".join(message_parts)
                
        except Exception as e:
            logger.error(f"Erreur export WhatsApp: {e}")
            return "❌ Erreur génération résumé"

# Test unitaire
if __name__ == "__main__":
    # Test basique
    exporter = RecapExporter()
    
    # Générer récapitulatif Excel
    excel_path = exporter.generate_complete_recap()
    if excel_path:
        print(f"✅ Récapitulatif Excel généré: {excel_path}")
    
    # Générer résumé PDF
    pdf_path = exporter.generate_pdf_summary()
    if pdf_path:
        print(f"✅ Résumé PDF généré: {pdf_path}")
    
    # Test export WhatsApp
    whatsapp_summary = exporter.export_for_whatsapp()
    print(f"\n📱 Résumé WhatsApp:\n{whatsapp_summary}")