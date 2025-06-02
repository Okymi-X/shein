# -*- coding: utf-8 -*-
"""
SHEIN_SEN - Gestionnaire de Données
Gestion des commandes, utilisateurs et exports Excel/PDF
"""

import json
import pandas as pd
import uuid
from datetime import datetime, timedelta
from typing import Dict, List, Optional, Any
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from loguru import logger

from config import Config

class DataManager:
    """Gestionnaire principal des données SHEIN_SEN"""
    
    def __init__(self):
        self.setup_logging()
        self.ensure_data_files()
    
    def setup_logging(self):
        """Configuration des logs"""
        logger.add(
            f"{Config.LOGS_DIR}/data_manager.log",
            rotation="1 day",
            retention="30 days",
            level="INFO"
        )
    
    def ensure_data_files(self):
        """S'assurer que les fichiers de données existent"""
        try:
            # Créer le fichier Excel des commandes s'il n'existe pas
            if not Path(Config.ORDERS_FILE).exists():
                self._create_orders_excel()
            
            # Créer le fichier JSON des utilisateurs s'il n'existe pas
            if not Path(Config.USERS_FILE).exists():
                self._create_users_json()
                
            logger.info("Fichiers de données initialisés")
            
        except Exception as e:
            logger.error(f"Erreur initialisation fichiers: {e}")
            raise
    
    def _create_orders_excel(self):
        """Créer le fichier Excel des commandes avec structure"""
        columns = [
            'order_id', 'user_phone', 'user_name', 'product_url', 
            'size', 'color', 'quantity', 'estimated_price', 
            'status', 'created_at', 'processed_at', 'notes'
        ]
        
        df = pd.DataFrame(columns=columns)
        
        with pd.ExcelWriter(Config.ORDERS_FILE, engine='openpyxl') as writer:
            # Feuille principale des commandes
            df.to_excel(writer, sheet_name='Commandes', index=False)
            
            # Feuille de résumé par utilisateur
            summary_df = pd.DataFrame(columns=[
                'user_phone', 'user_name', 'total_items', 'estimated_total', 'last_order'
            ])
            summary_df.to_excel(writer, sheet_name='Résumé_Utilisateurs', index=False)
            
            # Feuille de statistiques
            stats_df = pd.DataFrame(columns=[
                'metric', 'value', 'updated_at'
            ])
            stats_df.to_excel(writer, sheet_name='Statistiques', index=False)
        
        # Appliquer le formatage
        self._format_excel_file(Config.ORDERS_FILE)
        
        logger.info(f"Fichier Excel créé: {Config.ORDERS_FILE}")
    
    def _create_users_json(self):
        """Créer le fichier JSON des utilisateurs"""
        initial_data = {
            'users': {},
            'metadata': {
                'created_at': datetime.now().isoformat(),
                'last_updated': datetime.now().isoformat(),
                'total_users': 0
            }
        }
        
        with open(Config.USERS_FILE, 'w', encoding='utf-8') as f:
            json.dump(initial_data, f, indent=2, ensure_ascii=False)
        
        logger.info(f"Fichier utilisateurs créé: {Config.USERS_FILE}")
    
    def _format_excel_file(self, file_path: str):
        """Appliquer un formatage professionnel au fichier Excel"""
        try:
            wb = openpyxl.load_workbook(file_path)
            
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                
                # Style pour les en-têtes
                header_font = Font(bold=True, color="FFFFFF")
                header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                header_alignment = Alignment(horizontal="center", vertical="center")
                
                # Bordures
                thin_border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
                
                # Appliquer le style aux en-têtes (première ligne)
                if ws.max_row > 0:
                    for cell in ws[1]:
                        cell.font = header_font
                        cell.fill = header_fill
                        cell.alignment = header_alignment
                        cell.border = thin_border
                
                # Ajuster la largeur des colonnes
                for column in ws.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    
                    adjusted_width = min(max_length + 2, 50)
                    ws.column_dimensions[column_letter].width = adjusted_width
            
            wb.save(file_path)
            
        except Exception as e:
            logger.error(f"Erreur formatage Excel: {e}")
    
    def add_order(self, order_data: Dict) -> Optional[str]:
        """Ajouter une nouvelle commande"""
        try:
            # Générer un ID unique
            order_id = f"SHEIN_{datetime.now().strftime('%Y%m%d')}_{str(uuid.uuid4())[:8]}"
            
            # Préparer les données
            order_row = {
                'order_id': order_id,
                'user_phone': order_data.get('user_phone', ''),
                'user_name': self._get_user_name(order_data.get('user_phone', '')),
                'product_url': order_data.get('product_url', ''),
                'size': order_data.get('size', ''),
                'color': order_data.get('color', ''),
                'quantity': order_data.get('quantity', 1),
                'estimated_price': 0,  # À calculer plus tard
                'status': 'pending',
                'created_at': datetime.now().isoformat(),
                'processed_at': '',
                'notes': ''
            }
            
            # Lire le fichier Excel existant
            df = pd.read_excel(Config.ORDERS_FILE, sheet_name='Commandes')
            
            # Ajouter la nouvelle ligne
            new_df = pd.concat([df, pd.DataFrame([order_row])], ignore_index=True)
            
            # Sauvegarder
            with pd.ExcelWriter(Config.ORDERS_FILE, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
                new_df.to_excel(writer, sheet_name='Commandes', index=False)
            
            # Mettre à jour les résumés
            self._update_user_summary()
            self._update_statistics()
            
            # Reformater le fichier
            self._format_excel_file(Config.ORDERS_FILE)
            
            logger.info(f"Commande ajoutée: {order_id}")
            return order_id
            
        except Exception as e:
            logger.error(f"Erreur ajout commande: {e}")
            return None
    
    def get_user_orders(self, user_phone: str) -> List[Dict]:
        """Récupérer les commandes d'un utilisateur"""
        try:
            df = pd.read_excel(Config.ORDERS_FILE, sheet_name='Commandes')
            user_orders = df[df['user_phone'] == user_phone]
            
            return user_orders.to_dict('records')
            
        except Exception as e:
            logger.error(f"Erreur récupération commandes utilisateur: {e}")
            return []
    
    def get_all_orders(self, status: Optional[str] = None) -> List[Dict]:
        """Récupérer toutes les commandes avec filtre optionnel"""
        try:
            df = pd.read_excel(Config.ORDERS_FILE, sheet_name='Commandes')
            
            if status:
                df = df[df['status'] == status]
            
            return df.to_dict('records')
            
        except Exception as e:
            logger.error(f"Erreur récupération toutes commandes: {e}")
            return []
    
    def update_order_status(self, order_id: str, status: str, notes: str = '') -> bool:
        """Mettre à jour le statut d'une commande"""
        try:
            df = pd.read_excel(Config.ORDERS_FILE, sheet_name='Commandes')
            
            # Trouver et mettre à jour la commande
            mask = df['order_id'] == order_id
            if mask.any():
                df.loc[mask, 'status'] = status
                df.loc[mask, 'processed_at'] = datetime.now().isoformat()
                if notes:
                    df.loc[mask, 'notes'] = notes
                
                # Sauvegarder
                with pd.ExcelWriter(Config.ORDERS_FILE, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
                    df.to_excel(writer, sheet_name='Commandes', index=False)
                
                self._format_excel_file(Config.ORDERS_FILE)
                
                logger.info(f"Statut mis à jour: {order_id} -> {status}")
                return True
            
            return False
            
        except Exception as e:
            logger.error(f"Erreur mise à jour statut: {e}")
            return False
    
    def _get_user_name(self, user_phone: str) -> str:
        """Récupérer le nom d'un utilisateur"""
        try:
            with open(Config.USERS_FILE, 'r', encoding='utf-8') as f:
                data = json.load(f)
            
            user_info = data.get('users', {}).get(user_phone, {})
            return user_info.get('name', f'Utilisateur_{user_phone[-4:]}')
            
        except Exception as e:
            logger.error(f"Erreur récupération nom utilisateur: {e}")
            return f'Utilisateur_{user_phone[-4:] if user_phone else "Unknown"}'
    
    def _update_user_summary(self):
        """Mettre à jour le résumé par utilisateur"""
        try:
            df = pd.read_excel(Config.ORDERS_FILE, sheet_name='Commandes')
            
            # Grouper par utilisateur
            summary = df.groupby('user_phone').agg({
                'user_name': 'first',
                'quantity': 'sum',
                'estimated_price': 'sum',
                'created_at': 'max'
            }).reset_index()
            
            summary.columns = ['user_phone', 'user_name', 'total_items', 'estimated_total', 'last_order']
            
            # Sauvegarder dans la feuille résumé
            with pd.ExcelWriter(Config.ORDERS_FILE, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
                summary.to_excel(writer, sheet_name='Résumé_Utilisateurs', index=False)
            
        except Exception as e:
            logger.error(f"Erreur mise à jour résumé utilisateurs: {e}")
    
    def _update_statistics(self):
        """Mettre à jour les statistiques globales"""
        try:
            df = pd.read_excel(Config.ORDERS_FILE, sheet_name='Commandes')
            
            stats = [
                {'metric': 'Total Commandes', 'value': len(df), 'updated_at': datetime.now().isoformat()},
                {'metric': 'Total Utilisateurs', 'value': df['user_phone'].nunique(), 'updated_at': datetime.now().isoformat()},
                {'metric': 'Total Articles', 'value': df['quantity'].sum(), 'updated_at': datetime.now().isoformat()},
                {'metric': 'Commandes Pending', 'value': len(df[df['status'] == 'pending']), 'updated_at': datetime.now().isoformat()},
                {'metric': 'Commandes Completed', 'value': len(df[df['status'] == 'completed']), 'updated_at': datetime.now().isoformat()}
            ]
            
            stats_df = pd.DataFrame(stats)
            
            with pd.ExcelWriter(Config.ORDERS_FILE, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
                stats_df.to_excel(writer, sheet_name='Statistiques', index=False)
            
        except Exception as e:
            logger.error(f"Erreur mise à jour statistiques: {e}")
    
    def generate_user_summary(self, user_phone: str) -> str:
        """Générer un résumé textuel pour un utilisateur"""
        try:
            orders = self.get_user_orders(user_phone)
            
            if not orders:
                return "📋 Aucune commande trouvée."
            
            total_items = sum(order.get('quantity', 0) for order in orders)
            pending_count = len([o for o in orders if o.get('status') == 'pending'])
            completed_count = len([o for o in orders if o.get('status') == 'completed'])
            
            summary_parts = [
                f"📊 Résumé de vos commandes:",
                f"📦 Total articles: {total_items}",
                f"⏳ En attente: {pending_count}",
                f"✅ Complétées: {completed_count}",
                "",
                "📋 Dernières commandes:"
            ]
            
            # Afficher les 3 dernières commandes
            recent_orders = sorted(orders, key=lambda x: x.get('created_at', ''), reverse=True)[:3]
            
            for i, order in enumerate(recent_orders, 1):
                status_emoji = "⏳" if order.get('status') == 'pending' else "✅"
                size_info = f" - {order.get('size')}" if order.get('size') else ""
                color_info = f" - {order.get('color')}" if order.get('color') else ""
                
                summary_parts.append(
                    f"{status_emoji} {i}. Qté: {order.get('quantity', 1)}{size_info}{color_info}"
                )
            
            return "\n".join(summary_parts)
            
        except Exception as e:
            logger.error(f"Erreur génération résumé utilisateur: {e}")
            return "❌ Erreur lors de la génération du résumé."
    
    def get_statistics(self) -> Dict:
        """Récupérer les statistiques globales"""
        try:
            df = pd.read_excel(Config.ORDERS_FILE, sheet_name='Commandes')
            
            stats = {
                'total_orders': len(df),
                'total_users': df['user_phone'].nunique(),
                'total_items': int(df['quantity'].sum()),
                'pending_orders': len(df[df['status'] == 'pending']),
                'completed_orders': len(df[df['status'] == 'completed']),
                'last_updated': datetime.now().isoformat()
            }
            
            # Statistiques par jour (7 derniers jours)
            df['date'] = pd.to_datetime(df['created_at']).dt.date
            last_7_days = datetime.now().date() - timedelta(days=7)
            recent_df = df[df['date'] >= last_7_days]
            
            daily_stats = recent_df.groupby('date').size().to_dict()
            stats['daily_orders'] = {str(k): v for k, v in daily_stats.items()}
            
            return stats
            
        except Exception as e:
            logger.error(f"Erreur récupération statistiques: {e}")
            return {}
    
    def export_summary_pdf(self, output_path: Optional[str] = None) -> str:
        """Exporter un résumé en PDF"""
        try:
            from reportlab.lib.pagesizes import A4
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib import colors
            from reportlab.lib.units import inch
            
            if not output_path:
                output_path = f"{Config.DATA_DIR}/resume_shein_sen_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
            
            doc = SimpleDocTemplate(output_path, pagesize=A4)
            styles = getSampleStyleSheet()
            story = []
            
            # Titre
            title_style = ParagraphStyle(
                'CustomTitle',
                parent=styles['Heading1'],
                fontSize=24,
                spaceAfter=30,
                alignment=1  # Centré
            )
            
            story.append(Paragraph("SHEIN_SEN - Résumé des Commandes", title_style))
            story.append(Spacer(1, 20))
            
            # Statistiques globales
            stats = self.get_statistics()
            stats_data = [
                ['Métrique', 'Valeur'],
                ['Total Commandes', stats.get('total_orders', 0)],
                ['Total Utilisateurs', stats.get('total_users', 0)],
                ['Total Articles', stats.get('total_items', 0)],
                ['En Attente', stats.get('pending_orders', 0)],
                ['Complétées', stats.get('completed_orders', 0)]
            ]
            
            stats_table = Table(stats_data)
            stats_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 14),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            
            story.append(stats_table)
            story.append(Spacer(1, 30))
            
            # Tableau des commandes récentes
            df = pd.read_excel(Config.ORDERS_FILE, sheet_name='Commandes')
            recent_orders = df.tail(10)  # 10 dernières commandes
            
            if not recent_orders.empty:
                story.append(Paragraph("Dernières Commandes", styles['Heading2']))
                story.append(Spacer(1, 10))
                
                orders_data = [['ID', 'Utilisateur', 'Taille', 'Couleur', 'Qté', 'Statut']]
                
                for _, order in recent_orders.iterrows():
                    orders_data.append([
                        str(order.get('order_id', ''))[:15] + '...',
                        str(order.get('user_name', ''))[:15],
                        str(order.get('size', '')),
                        str(order.get('color', ''))[:10],
                        str(order.get('quantity', '')),
                        str(order.get('status', ''))
                    ])
                
                orders_table = Table(orders_data)
                orders_table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                    ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, 0), 10),
                    ('FONTSIZE', (0, 1), (-1, -1), 8),
                    ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                    ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                    ('GRID', (0, 0), (-1, -1), 1, colors.black)
                ]))
                
                story.append(orders_table)
            
            # Générer le PDF
            doc.build(story)
            
            logger.info(f"PDF généré: {output_path}")
            return output_path
            
        except Exception as e:
            logger.error(f"Erreur génération PDF: {e}")
            return ""

# Test unitaire
if __name__ == "__main__":
    # Test basique
    dm = DataManager()
    
    # Test ajout commande
    test_order = {
        'user_phone': '+221701234567',
        'product_url': 'https://www.shein.com/fr/item12345',
        'size': 'M',
        'color': 'Rouge',
        'quantity': 2
    }
    
    order_id = dm.add_order(test_order)
    print(f"Commande ajoutée: {order_id}")
    
    # Test récupération
    orders = dm.get_user_orders('+221701234567')
    print(f"Commandes utilisateur: {len(orders)}")
    
    # Test statistiques
    stats = dm.get_statistics()
    print(f"Statistiques: {stats}")